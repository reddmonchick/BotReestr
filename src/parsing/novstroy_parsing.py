from src.dto import FiltersParser

from curl_cffi import requests
from loguru import logger
from datetime import datetime
from openpyxl import Workbook, load_workbook
import time

class Novstroy:
    def __init__(self, filters: FiltersParser):
        self.filters = filters
        self.session = requests.Session()

    def create_excel_file(self, filterdto: FiltersParser):
        pass

    def is_date_in_range(self, input_date_str, start_date_str, end_date_str):
        # Parse the input date in the format %d.%m.%Y
        input_date = datetime.strptime(input_date_str, "%d.%m.%Y")

        # Parse the start and end dates in the format %Y-%m-%d
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")

        # Check if the input date is within the range
        return start_date <= input_date <= end_date

    def parse(self):
        filterdto = self.filters

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
            # 'Accept-Encoding': 'gzip, deflate, br, zstd',
            'Content-Type': 'application/json',
            'Origin': 'https://reestr.nopriz.ru',
            'Connection': 'keep-alive',
            'Referer': 'https://reestr.nopriz.ru/',
            # 'Cookie': 'BITRIX_SM_GUEST_ID=2592147; BITRIX_SM_LAST_VISIT=17.05.2024%2023%3A24%3A43; BITRIX_SM_LAST_ADV=5_Y; _ym_uid=170358633291607385; _ym_d=1712579416; _ym_isad=1',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'Priority': 'u=1',
        }
        json_data = {
            'filters': {
                'registry_registration_date': [
                ],
                'suspension_date': [
                ],
            },
            'page': 1,
            'pageCount': '5000',
            'sortBy': {
                'registry_registration_date': 'DESC',
            },
        }

        column_xlsx = ['Название компании', 'ИНН', 'Номера', 'Дата внесения в реестр', 'Статус']
        flattens_xlsx = []

        if filterdto.date_join:
            column_xlsx[-1] = 'Дата внесения в реестр'
            json_data['filters']['registry_registration_date'].extend([filterdto.data_start, filterdto.data_end])
        else:
            column_xlsx[-1] = 'Дата прекращения членства'
            json_data['filters']['suspension_date'].extend([filterdto.data_start, filterdto.data_end])

        if filterdto.status is None:
            pass
        elif filterdto.status == False:
            json_data['filters']['member_status'] = 2
        elif filterdto.status == True:
            json_data['filters']['member_status'] = 1
        print(filterdto)

        break_count = 0
        while True:
            try:
                response = self.session.post('https://reestr.nostroy.ru/api/sro/all/member/list', headers=headers,
                                             json=json_data, verify=False).json()
                break
            except Exception as ex:
                logger.error(f'Ошибка в получении количество страниц у novstroy {ex}')
                break_count += 1
                if break_count >= 10:
                    response = {}
                    break

        col_pages = int(response.get('data', {}).get('countPages', 0))
        logger.debug(f'Количество страниц novstroy {col_pages}')
        id_companys = []
        for page in range(1, col_pages + 1):
            break_count = 0
            while True:
                try:
                    json_data['page'] = page
                    response = self.session.post('https://reestr.nostroy.ru/api/sro/all/member/list', headers=headers, json=json_data, verify=False)
                    resp_json = response.json()
                    result = resp_json.get('data', []).get('data', {})
                    logger.debug(f'Парсинг компаний novstroy {response.status_code} {page} | {col_pages}')
                    for res in result:
                        id_companys.append(res.get('id', 0))
                    time.sleep(5)
                    break
                except Exception as ex:
                    logger.error(f'Ошибка в получении страниц у novstroy {ex} {response.status_code}')
                    break_count += 1
                    if break_count >= 10:
                        break

        count_id = 0
        for id_company in id_companys:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0',
                'Accept': 'application/json, text/plain, */*',
                'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
                # 'Accept-Encoding': 'gzip, deflate, br, zstd',
                # Already added when you pass json=
                # 'Content-Type': 'application/json',
                'Origin': 'https://reestr.nopriz.ru',
                'Connection': 'keep-alive',
                'Referer': 'https://reestr.nopriz.ru/member/19516447',
                # 'Cookie': 'BITRIX_SM_GUEST_ID=2592147; BITRIX_SM_LAST_VISIT=17.05.2024%2023%3A24%3A43; BITRIX_SM_LAST_ADV=5_Y; _ym_uid=170358633291607385; _ym_d=1712579416',
                'Sec-Fetch-Dest': 'empty',
                'Sec-Fetch-Mode': 'cors',
                'Sec-Fetch-Site': 'same-origin',
            }
            break_count = 0
            while True:
                try:
                    response = self.session.post(f'https://reestr.nostroy.ru/api/member/{id_company}/info',
                                                 headers=headers, verify=False)
                    logger.debug(f'Парсинг реестра компании novstroy {response.status_code} {id_company} {count_id} | {len(id_companys)} ')
                    result = response.json().get('data', {})
                    name_company = result.get('short_description', '')
                    inn = result.get('inn', '')
                    number = result.get('phones', '')
                    date_sro = ''
                    if self.filters.date_join:
                        date_sro = result.get('registry_registration_date', '')
                        if date_sro:
                            date_object = datetime.fromisoformat(date_sro)
                            # Форматируем дату в нужный формат
                            date_sro = date_object.strftime("%d.%m.%Y")
                        else:
                            break
                    else:
                        date_sro = result.get('suspension_date', '')
                        if date_sro:
                            date_object = datetime.fromisoformat(date_sro)
                            # Форматируем дату в нужный формат
                            date_sro = date_object.strftime("%d.%m.%Y")
                            res = self.is_date_in_range(date_sro, filterdto.data_start, filterdto.data_end)
                            if not res:
                                break
                        else:
                            break
                    status_sro = result.get('member_status', {}).get('title')
                    flattens_xlsx.append([name_company, inn, number, date_sro, status_sro])
                    count_id += 1
                    break
                except Exception as ex:
                    logger.error(f'Ошибка в получении сро у novstroy {ex} {response.status_code} {id_company}')
                    break_count += 1
                    if break_count >= 10:
                        break

        file_path_write = f'{self.filters.user_id}_novstroy.xlsx'
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(column_xlsx)
        workbook.save(file_path_write)

        workbook = load_workbook(file_path_write)
        worksheet = workbook.active
        for rrik in flattens_xlsx:
            worksheet.append(rrik)
        workbook.save(file_path_write)
        logger.info(f'Сохранили в {file_path_write} результат парсинга nopriz')